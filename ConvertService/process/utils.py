import datetime
import os
import re
import csv
import json
import fitz
import xml.etree.ElementTree as ET
from pathlib import Path
import pandas as pd
import jaconv
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
import logging
from .fetch_data import RuleFixedID, FixedValueFetcher

logger = logging.getLogger(__name__)


class DisplayData:
    @staticmethod
    def filter_list(headers, hidden_headers, data):
        """
        Filter headers and data based on visible headers.
        """
        try:
            visible_indices = [i for i, header in enumerate(headers) if header not in hidden_headers]
            filtered_headers = [header for i, header in enumerate(headers) if i in visible_indices]

            filtered_data = []
            for row in data:
                if len(row) < len(headers):
                    logger.warning(f"Skipping row due to mismatched length: {row}")
                    continue
                try:
                    filtered_row = [row[i] if i < len(row) else "" for i in visible_indices]
                    filtered_data.append(filtered_row)
                except IndexError as e:
                    logger.error(f"Error accessing row: {row}, error: {e}")
                    continue

            return filtered_headers, filtered_data
        except Exception as e:
            logger.error(f"Error filtering data: {e}")
            return [], []

    @staticmethod
    def get_list_data(redis_client, keys, hidden_formatted_header):
        try:
            visible_indices = [
                header.get('index_value') for header in hidden_formatted_header
                if header.get('index_value', False)
            ]

            result = []
            max_workers = os.cpu_count() or 4

            for key in keys:
                try:
                    raw_data = redis_client.get(key)
                    if raw_data:
                        data = json.loads(raw_data.decode('utf-8'))
                        with ThreadPoolExecutor(max_workers=max_workers) as executor:
                            filtered_rows = []
                            futures = []

                            def remove_columns(row):
                                indices_to_remove = sorted(visible_indices, reverse=True)
                                new_row = list(row)
                                for idx in indices_to_remove:
                                    if idx < len(new_row):
                                        new_row.pop(idx)
                                return new_row

                            for row in data:
                                futures.append(
                                    executor.submit(remove_columns, row)
                                )

                            for future in futures:
                                try:
                                    filtered_row = future.result()
                                    if filtered_row:
                                        filtered_rows.append(filtered_row)
                                except Exception as e:
                                    logger.error(f"Error processing row: {e}")
                                    continue

                        if filtered_rows:
                            result.append({
                                "data": filtered_rows,
                                "key": key.decode('utf-8')
                            })
                except Exception as e:
                    logger.error(f"Error reading formatted data from Redis key {key}: {e}")
                    return []

            return result
        except Exception as e:
            logger.error(f"Error combining formatted data: {e}")
            return []

    @staticmethod
    def get_paginated_data(redis_client, keys, hidden_formatted_header, page=1, page_size=20):
        try:
            visible_indices = [
                header.get('index_value') for header in hidden_formatted_header
                if header.get('index_value', False)
            ]

            result = []

            start_idx = (page - 1) * page_size
            end_idx = start_idx + page_size

            processed_rows = 0
            total_rows = 0

            for key in keys:
                try:
                    raw_data = redis_client.get(key)
                    if not raw_data:
                        continue

                    data = json.loads(raw_data.decode('utf-8'))
                    data_rows_count = len(data)
                    total_rows += data_rows_count

                    if processed_rows + data_rows_count <= start_idx:
                        processed_rows += data_rows_count
                        continue

                    local_start = max(0, start_idx - processed_rows)
                    local_end = min(data_rows_count, end_idx - processed_rows)

                    current_page_data = data[local_start:local_end]

                    filtered_rows = []
                    for row in current_page_data:
                        if isinstance(row, list):
                            new_row = list(row)
                            for idx in sorted(visible_indices, reverse=True):
                                if idx < len(new_row):
                                    new_row.pop(idx)
                            filtered_rows.append(new_row)

                    if filtered_rows:
                        result.append({
                            "data": filtered_rows,
                            "key": key.decode('utf-8')
                        })

                    processed_rows += data_rows_count
                    if processed_rows >= end_idx:
                        break

                except Exception as e:
                    logger.error(f"Reading Rdis key failed {key}: {e}")
                    continue

            return result, total_rows

        except Exception as e:
            logger.error(f"Failed to get pagination: {e}")
            return [], 0


class FileFormatMapper:
    FORMAT_MAPPING = {
        'CSV_C_SJIS': {'delimiter': ',', 'encoding': 'shift_jis'},
        'CSV_C_UTF-8': {'delimiter': ',', 'encoding': 'utf-8'},
        'CSV_T_SJIS': {'delimiter': '\t', 'encoding': 'shift_jis'},
        'CSV_T_UTF-8': {'delimiter': '\t', 'encoding': 'utf-8'},
    }

    @staticmethod
    def get_format_details(file_format_id: str):
        return FileFormatMapper.FORMAT_MAPPING.get(file_format_id, None)


class FileProcessor:
    Fallback_Encodings = ['utf-8', 'shift_jis', 'cp932', 'iso-8859-1']

    @staticmethod
    def process_file(file_path, headers):
        file_path = Path(file_path)
        if file_path.suffix.lower() in ['.csv', '.tsv', '.txt']:
            return FileProcessor.process_csv(file_path, headers)
        elif file_path.suffix.lower() in ['.json']:
            return FileProcessor.process_json(file_path, headers)
        elif file_path.suffix.lower() in ['.xml']:
            return FileProcessor.process_xml(file_path, headers)
        elif file_path.suffix.lower() in ['.xlsx', '.xls']:
            return FileProcessor.process_excel(file_path, headers)
        elif file_path.suffix.lower() in ['.pdf']:
            return FileProcessor.process_pdf(file_path, headers)
        else:
            raise ValueError(f"Unsupported file type: {file_path.suffix}")

    @staticmethod
    def process_csv(file_path, headers):
        """
        Process CSV files with fallback for different encodings
        """
        try:
            encodings = ['utf-8', 'shift_jis', 'cp932', 'iso-8859-1']

            with open(file_path, 'r', encoding=encodings[0], errors='ignore') as f:
                first_line = f.readline().strip()

            delimiters = [',', ';', '\t', '|']
            delimiter_counts = {d: first_line.count(d) for d in delimiters}
            delimiter = max(delimiter_counts.items(), key=lambda x: x[1])[0]

            try:
                return FileProcessor._read_csv(file_path, headers, delimiter, encodings[0])
            except UnicodeDecodeError:
                pass

            for encoding in encodings[1:]:
                try:
                    return FileProcessor._read_csv(file_path, headers, delimiter, encoding)
                except UnicodeDecodeError:
                    continue

            raise ValueError("Failed to process CSV file with all attempted encodings.")
        except Exception as e:
            logger.error(f"Error processing CSV file {file_path}: {e}")
            raise

    @staticmethod
    def _read_csv(file_path, headers, delimiter, encoding):
        """
        Helper method to read CSV files
        """
        try:
            with open(file_path, 'r', encoding=encoding, newline='') as csvfile:
                reader = csv.DictReader(csvfile, delimiter=delimiter)
                data = []
                for row_idx, row in enumerate(reader):
                    try:
                        filtered_row = {header: row.get(header, "") for header in headers}
                        data.append(filtered_row)
                    except Exception as row_error:
                        logger.warning(f"Error processing row {row_idx + 1}: {row_error}. Skipping this row.")
                        continue

            logger.info(f"Successfully read CSV with encoding {encoding}.")
            return data
        except UnicodeDecodeError as e:
            logger.warning(f"Failed to read CSV with encoding {encoding}: {e}")
            raise

    @staticmethod
    def process_excel(file_path, headers):
        """
        Process Excel files (both .xls and .xlsx formats)
        """
        try:
            file_extension = os.path.splitext(file_path)[1].lower()
            if file_extension == '.xls':
                try:
                    import xlrd
                    workbook = xlrd.open_workbook(file_path)
                    sheet = workbook.sheet_by_index(0)

                    header_row = [str(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
                    header_indices = {header: idx for idx, header in enumerate(header_row) if header in headers}

                    data = []
                    for row_idx in range(1, sheet.nrows):
                        try:
                            if all(sheet.cell_value(row_idx, col) == '' for col in range(sheet.ncols)):
                                continue

                            row_data = {}
                            for header in headers:
                                if header in header_indices:
                                    cell_value = sheet.cell_value(row_idx, header_indices[header])
                                    if sheet.cell_type(row_idx, header_indices[header]) == xlrd.XL_CELL_DATE:
                                        datetime_obj = xlrd.xldate_as_datetime(cell_value, workbook.datemode)
                                        cell_value = datetime_obj.strftime("%Y/%m/%d")
                                    row_data[header] = str(cell_value) if cell_value is not None else ""
                                else:
                                    row_data[header] = ""
                            data.append(row_data)
                        except Exception as e:
                            logger.error(f"Error processing row {row_idx}: {e}")
                            continue

                    return data
                except ImportError:
                    logger.error("xlrd library not installed. Cannot process .xls files.")
                    raise ImportError("xlrd library required for processing .xls files")
            else:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                sheet = workbook.active

                first_row = list(sheet.iter_rows(min_row=1, max_row=1))[0]
                header_row = [str(cell.value).strip() if cell.value else "" for cell in first_row]
                header_indices = {header: idx for idx, header in enumerate(header_row) if header in headers}

                data = []
                for row in list(sheet.iter_rows())[1:]:  # Skip header row
                    try:
                        if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                            continue

                        row_data = {}
                        for header in headers:
                            if header in header_indices and header_indices[header] < len(row):
                                cell_value = row[header_indices[header]].value

                                if isinstance(cell_value, datetime.datetime) or isinstance(cell_value, datetime.date):
                                    cell_value = cell_value.strftime("%Y/%m/%d")

                                row_data[header] = str(cell_value) if cell_value is not None else ""
                            else:
                                row_data[header] = ""
                        data.append(row_data)
                    except Exception as e:
                        logger.error(f"Error processing row: {e}")
                        continue

                return data

        except Exception as e:
            logger.error(f"Error processing Excel file {file_path}: {e}")
            return []

    @staticmethod
    def process_json(file_path, headers):
        """
        Process JSON files with fallback for different encodings
        """
        try:
            for encoding in ['utf-8', 'shift_jis', 'cp932', 'iso-8859-1']:
                try:
                    with open(file_path, 'r', encoding=encoding) as jsonfile:
                        json_data = json.load(jsonfile)
                        if isinstance(json_data, list):
                            data = []
                            for item in json_data:
                                data.append({header: item.get(header, "") for header in headers})
                        elif isinstance(json_data, dict):
                            data = [{header: json_data.get(header, "") for header in headers}]
                        else:
                            data = []
                        return data
                except (json.JSONDecodeError, UnicodeDecodeError) as e:
                    logger.warning(f"Error decoding JSON with encoding {encoding}: {e}")
            raise ValueError("Failed to process JSON file with all attempted encodings.")
        except Exception as e:
            logger.error(f"Error processing JSON file {file_path}: {e}")
            raise

    @staticmethod
    def process_xml(file_path, headers):
        """
        Process XML files with fallback for different encodings
        """
        try:
            for encoding in ['utf-8', 'shift_jis', 'cp932', 'iso-8859-1']:
                try:
                    with open(file_path, 'r', encoding=encoding) as xmlfile:
                        tree = ET.parse(xmlfile)
                        root = tree.getroot()
                        data = []
                        for elem in root.findall('.//record'):
                            data.append(
                                {header: elem.find(header).text if elem.find(header) is not None else "" for header in
                                 headers})
                        return data
                except (ET.ParseError, UnicodeDecodeError) as e:
                    logger.warning(f"Error parsing XML with encoding {encoding}: {e}")
            raise ValueError("Failed to process XML file with all attempted encodings.")
        except Exception as e:
            logger.error(f"Error processing XML file {file_path}: {e}")
            raise

    @staticmethod
    def process_pdf(file_path, headers):
        """
        Process PDF files with structured data extraction
        """
        try:
            pdf_document = fitz.open(file_path)
            data = []
            for page in pdf_document:
                form_fields = page.widgets()
                if form_fields:
                    row_data = {}
                    for widget in form_fields:
                        field_name = widget.field_name
                        field_value = widget.field_value or ""
                        if field_name in headers:
                            row_data[field_name] = field_value
                        else:
                            row_data[field_name] = ""
                    data.append(row_data)
            pdf_document.close()
            return data
        except Exception as e:
            logger.error(f"Error processing PDF file {file_path}: {e}")
            raise


class DataFormatter:
    """
    A class to perform various data transformations and file formatting tasks.
    """
    class RuleFixedMapping(RuleFixedID):
        pass

    RULE_MAPPING = {
        "CR_NOT_CHANGE": lambda value: value,
        "CR_DATE1": lambda value: DataFormatter.convert_date(value, '%Y/%m/%d'),
        "CR_DATE2": lambda value: DataFormatter.convert_date(value, '%Y-%m-%d'),
        "CR_G_12": lambda value: DataFormatter.convert_gender(value, '12'),
        "CR_G_MF": lambda value: DataFormatter.convert_gender(value, 'MF'),
        "CR_KANA_F-H": lambda value: DataFormatter.convert_kana(value, 'full_to_half'),
        "CR_KANA_H-F": lambda value: DataFormatter.convert_kana(value, 'half_to_full'),
        "CR_POSTAL_FORMAT": lambda value: DataFormatter.convert_postal_code(value),
        "CR_TIME": lambda value: DataFormatter.convert_time(value),
    }

    @staticmethod
    def apply_rule(value, rule_id):
        try:
            rule_function = DataFormatter.RULE_MAPPING.get(rule_id, lambda v: v)
            transformed_value = rule_function(value)
            return transformed_value
        except Exception as e:
            logger.error(f"Error applying rule {rule_id} to value {value}: {e}")
            return value

    @staticmethod
    def is_fixed_rule(rule_id):
        return rule_id in DataFormatter.RuleFixedMapping.get_values()

    @staticmethod
    def format_data_with_rules(row, rules, before_headers, after_headers, tenant_id, data_format_id=None):
        try:
            mapped_row = [""] * len(after_headers)

            if isinstance(row, dict):
                index_to_header = {h['index_value']: h['header_name'] for h in before_headers}

                for rule_id, idx_before, idx_after in rules:
                    try:
                        if idx_after < len(mapped_row):
                            header_name = index_to_header.get(idx_before)
                            if header_name and header_name in row:
                                value = row[header_name]
                                mapped_row[idx_after] = value
                            else:
                                logger.info(f"Header name {header_name} not found in row or is None")
                    except Exception as e:
                        logger.error(f"Error in rule mapping: {e}")

            elif isinstance(row, list):
                before_indices = [h['index_value'] for h in before_headers]
                for rule_id, idx_before, idx_after in rules:
                    try:
                        if idx_after < len(mapped_row):
                            if idx_before < len(row):
                                value = row[idx_before]
                                if idx_before in before_indices:
                                    mapped_row[idx_after] = value
                                else:
                                    logger.info(f"Index {idx_before} not in before_indices")
                            else:
                                logger.info(f"Index {idx_before} out of range for row")
                    except Exception as e:
                        logger.error(f"Error in rule mapping: {e}")

            else:
                logger.error(f"Unsupported row type: {type(row)}")
                return [""] * len(after_headers)

            for rule_id, idx_before, idx_after in rules:
                try:
                    if idx_after < len(mapped_row):
                        before_value = mapped_row[idx_after]
                        if DataFormatter.is_fixed_rule(rule_id):
                            mapped_row[idx_after] = DataFormatter.convert_fixed_value(
                                before_value, rule_id, tenant_id, data_format_id
                            )
                        else:
                            mapped_row[idx_after] = DataFormatter.apply_rule(
                                before_value, rule_id
                            )
                        after_value = mapped_row[idx_after]

                except Exception as e:
                    logger.error(f"Error applying rule: {e}")
            return mapped_row

        except Exception as e:
            logger.error(f"Error in format_data_with_rules: {e}", exc_info=True)
            return [""] * len(after_headers)

    @staticmethod
    def convert_fixed_value(value, rule_id, tenant_id, data_format_id=None):
        try:
            after_value = FixedValueFetcher.get_value_mapping(tenant_id, rule_id, value, data_format_id)

            return after_value if after_value else value

        except Exception as e:
            logger.error(f"Error in convert_fixed_value for rule {rule_id}, tenant {tenant_id}: {value} -> {e}")
            return value

    @staticmethod
    def convert_date(value, target_format='%Y/%m/%d'):
        """
        Convert a date string to the desired format.
        Supported formats:
        - Japanese Era formats
        - ISO formats (YYYY/MM/DD, YYYY-MM-DD)
        - Kanji dates (YYYY年MM月DD日)
        - US-style formats (MM/DD/YYYY, MM-DD-YYYY)
        - Custom formats like DD/MM/YYYY, YYYY\DD\MM

        :param value: The date value to be converted
        :param target_format: The desired output format
        :return: Formatted date string or original value if conversion fails
        """
        try:
            if not value or not isinstance(value, str):
                return ""

            value = value.strip()
            date_result = None

            # 1. Handle Japanese Era formats
            era_map = {"S": 1925, "H": 1988, "R": 2018}
            era_match = re.match(r'([S|H|R])\s*(\d{1,2})[.\-年](\d{1,2})[.\-月](\d{1,2})日?', value)
            if era_match:
                era, year, month, day = era_match.groups()
                base_year = era_map.get(era[0], 0)
                year = int(year) + base_year
                date_result = datetime.datetime(year, int(month), int(day))

            kanji_era_match = re.match(r'(昭和|平成|令和)\s*(\d{1,2})[.\-年](\d{1,2})[.\-月](\d{1,2})日?', value)
            if kanji_era_match:
                kanji_era, year, month, day = kanji_era_match.groups()
                era = {"昭和": 1925, "平成": 1988, "令和": 2018}[kanji_era]
                year = int(year) + era
                date_result = datetime.datetime(year, int(month), int(day))

            # 2. Handle ISO-like formats
            if not date_result:
                iso_match = re.match(r'(\d{4})[\/\-](\d{1,2})[\/\-](\d{1,2})', value)
                if iso_match:
                    year, month, day = map(int, iso_match.groups())
                    date_result = datetime.datetime(year, month, day)

            # 3. Handle Kanji dates
            if not date_result:
                kanji_match = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', value)
                if kanji_match:
                    year, month, day = map(int, kanji_match.groups())
                    date_result = datetime.datetime(year, month, day)

            # 4. Handle US-style formats with time
            if not date_result:
                us_match_with_time = re.match(
                    r'(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})\s(\d{1,2}):(\d{1,2}):(\d{1,2})\s(AM|PM)', value)
                if us_match_with_time:
                    month, day, year, hour, minute, second, period = us_match_with_time.groups()
                    hour = int(hour)
                    if period == 'PM' and hour != 12:
                        hour += 12
                    elif period == 'AM' and hour == 12:
                        hour = 0
                    date_result = datetime.datetime(int(year), int(month), int(day), hour, int(minute), int(second))

            # 5. Handle US-style formats without time (e.g., MM/DD/YYYY or MM-DD-YYYY)
            if not date_result:
                us_match = re.match(r'(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})', value)
                if us_match:
                    month, day, year = map(int, us_match.groups())
                    date_result = datetime.datetime(year, month, day)

            # 6. Handle DD/MM/YYYY or YYYY\DD\MM
            if not date_result:
                custom_match = re.match(r'(\d{1,2})[\/\-\\](\d{1,2})[\/\-\\](\d{4})', value)
                if custom_match:
                    day, month, year = map(int, custom_match.groups())
                    date_result = datetime.datetime(year, month, day)

            if not date_result:
                reverse_match = re.match(r'(\d{4})[\/\-\\](\d{1,2})[\/\-\\](\d{1,2})', value)
                if reverse_match:
                    year, day, month = map(int, reverse_match.groups())
                    date_result = datetime.datetime(year, month, day)

            # 7. Convert to the desired format
            if date_result:
                if target_format == '%d/%m/%Y':
                    return date_result.strftime('%d/%m/%Y')
                return date_result.strftime(target_format)

            return value
        except Exception as e:
            logger.error(f"Error converting date: {value} -> {e}")
            return value

    @staticmethod
    def convert_gender(value, gender_type='12'):
        """
        Convert gender format based on the specified type.

        Supported types:
        - '12': Male = 1, Female = 2
        - 'MF': Male = M, Female = F
        - 'KANJI': Male = 男, Female = 女
        - 'KANJI_FULL': Male = 男性, Female = 女性
        """
        try:
            gender_map = {
                '12': {"男": "1", "女": "2", "男性": "1", "女性": "2", "1": "1", "2": "2"},
                'MF': {"男": "M", "女": "F", "男性": "M", "女性": "F", "M": "M", "F": "F"},
                'KANJI': {"1": "男", "2": "女", "M": "男", "F": "女", "男性": "男", "女性": "女"},
                'KANJI_FULL': {"1": "男性", "2": "女性", "M": "男性", "F": "女性", "男": "男性", "女": "女性"},
            }

            return gender_map.get(gender_type, {}).get(value, value)
        except Exception as e:
            logger.error(f"Error in convert_gender: {e}")
            return ""

    @staticmethod
    def convert_kana(value, kana_type='full_to_half'):
        """
        Convert Kana between full-width and half-width.
        - 'full_to_half': Convert full-width Kana to half-width.
        - 'half_to_full': Convert half-width Kana to full-width.
        """
        try:
            if not value:
                return ""

            value = jaconv.hira2kata(value)

            if kana_type == 'full_to_half':
                result = jaconv.z2h(value, kana=True)
                result = result.replace('\u3000', ' ')
                return result
            elif kana_type == 'half_to_full':
                result = jaconv.h2z(value, kana=True)
                result = result.replace(' ', '\u3000')
                return result
            else:
                logger.warning(f"Unsupported kana type: {kana_type}")
                return value
        except Exception as e:
            logger.error(f"Error in convert_kana: {e}")
            return ""

    @staticmethod
    def convert_postal_code(value):
        """
        Automatically detects and converts various postal code formats to a standard format (XXX-XXXX).

        :param value: The postal code value to be converted
        :return: Standardized postal code or original value if conversion fails
        """
        try:
            if not value or not isinstance(value, str):
                return ""

            value = value.strip()
            value = re.sub(r"[^\d\-]", "", value)

            postal_match = re.match(r"(\d{3})[-\s]?(\d{4})", value)
            if postal_match:
                return f"{postal_match.group(1)}-{postal_match.group(2)}"

            continuous_digits_match = re.match(r"(\d{3})(\d{4})", value)
            if continuous_digits_match:
                return f"{continuous_digits_match.group(1)}-{continuous_digits_match.group(2)}"

            return value
        except Exception as e:
            logger.error(f"Error converting postal code: {value} -> {e}")
            return value

    @staticmethod
    def convert_time(value):
        """
        Convert time from hh:mm to custom format:
        - hh:01 -> hh:15 => hh15
        - hh:16 -> hh:30 => hh30
        - hh:31 -> hh:45 => hh45
        - hh:46 -> hh:00 => hh00

        :param value: The time value to be converted
        :return: Formatted time or original value if conversion fails
        """
        try:
            if not value or not isinstance(value, str):
                return ""

            value = value.strip()
            time_match = re.match(r"(\d{1,2}):(\d{2})", value)
            if not time_match:
                return value

            hour, minute = map(int, time_match.groups())

            if 1 <= minute <= 15:
                return f"{hour:02d}15"
            elif 16 <= minute <= 30:
                return f"{hour:02d}30"
            elif 31 <= minute <= 45:
                return f"{hour:02d}45"
            elif 46 <= minute <= 59 or minute == 0:
                next_hour = (hour + 1) % 24 if minute >= 46 else hour
                return f"{next_hour:02d}00"

            return value
        except Exception as e:
            logger.error(f"Error converting time: {value} -> {e}")
            return value


    @staticmethod
    def write_dict(data, file_paths):

        results = {}
        try:
            for index, (key, path) in enumerate(file_paths.items()):
                try:
                    with open(path, 'w', encoding='utf-8') as jsonfile:
                        json.dump(data[index], jsonfile, ensure_ascii=False, indent=4)
                    results[key] = f"Successfully created file at {path}"
                except Exception as e:
                    logger.error(f"Error writing to file {path}: {e}")
                    results[key] = f"Error: {str(e)}"
            return results
        except Exception as e:
            logger.error(f"Error in write_dict: {e}")
            return {"error": str(e)}

    @staticmethod
    def write_csv(data, file_path, delimiter=',', encoding='utf-8'):
        """
        Write data to a CSV file with proper line endings based on the platform.
        """
        try:
            with open(file_path, 'w', newline='', encoding=encoding) as csvfile:
                writer = csv.writer(csvfile, delimiter=delimiter)
                writer.writerows(data)
            logger.info(f"CSV file successfully created: {file_path}")
            return f"CSV file successfully created: {file_path}"
        except Exception as e:
            logger.error(f"Error writing CSV file: {file_path} - {e}")
            return f"CSV file creation error: {e}"

    @staticmethod
    def write_json(data, file_path):
        """
        Write data to a JSON file.
        """
        try:
            with open(file_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(data, jsonfile, ensure_ascii=False, indent=4)
            logger.info(f"JSON file successfully created: {file_path}")
            return f"JSON file successfully created: {file_path}"
        except Exception as e:
            logger.error(f"Error writing JSON file: {file_path} - {e}")
            return f"JSON file creation error: {e}"

    @staticmethod
    def write_xml(data, file_path):
        """
        Write data to an XML file.
        """
        try:
            root = ET.Element("root")
            for row in data:
                item = ET.SubElement(root, "record")
                for i, value in enumerate(row):
                    ET.SubElement(item, f"field{i+1}").text = value

            tree = ET.ElementTree(root)
            tree.write(file_path, encoding="utf-8", xml_declaration=True)
            logger.info(f"XML file successfully created: {file_path}")
            return f"XML file successfully created: {file_path}"
        except Exception as e:
            logger.error(f"Error writing XML file: {file_path} - {e}")
            return f"XML file creation error: {e}"

    @staticmethod
    def format_file(data, file_format, file_path):
        """
        Format file output based on the specified format.
        Supported formats: CSV, JSON, XML.
        """
        try:
            if file_format == 'CSV_C_SJIS':
                return DataFormatter.write_csv(data, file_path, delimiter=',', encoding='shift_jis')
            elif file_format == 'CSV_C_UTF-8':
                return DataFormatter.write_csv(data, file_path, delimiter=',', encoding='utf-8')
            elif file_format == 'CSV_T_SJIS':
                return DataFormatter.write_csv(data, file_path, delimiter='\t', encoding='shift_jis')
            elif file_format == 'CSV_T_UTF-8':
                return DataFormatter.write_csv(data, file_path, delimiter='\t', encoding='utf-8')
            elif file_format == 'JSON':
                return DataFormatter.write_json(data, file_path)
            elif file_format == 'XML':
                return DataFormatter.write_xml(data, file_path)
            else:
                logger.error(f"Unsupported file format: {file_format}")
                return f"Unsupported file format: {file_format}"
        except Exception as e:
            logger.error(f"Error formatting file: {file_path} - {e}")
            return f"File formatting error: {e}"


class ProcessHeader:

    @staticmethod
    def get_csv_header(file, delimiter, encoding):
        """
        Get headers from CSV file with support for different line endings.
        """
        try:
            file.seek(0)
            content = file.read().decode(encoding, errors='replace')
            content = content.replace('\r\n', '\n').replace('\r', '\n')

            if not delimiter:
                first_line = content.splitlines()[0] if content.splitlines() else ""
                delimiters = [',', ';', '\t', '|']
                delimiter_counts = {d: first_line.count(d) for d in delimiters}

                if any(count > 0 for count in delimiter_counts.values()):
                    delimiter = max(delimiter_counts.items(), key=lambda x: x[1])[0]
                    logger.info(f"Auto-detected delimiter: {delimiter}")
                else:
                    delimiter = ','
                    logger.warning(f"No delimiter found in CSV, using default: {delimiter}")

            reader = csv.reader(content.splitlines(), delimiter=delimiter)
            try:
                header = next(reader)
                return [col.strip() for col in header]
            except StopIteration:
                logger.warning("Empty CSV file or no header found")
                return []
        except Exception as e:
            logger.error(f"Error reading CSV header: {e}")
            return []

    @staticmethod
    def get_pdf_header(file):
        try:
            file.seek(0)
            reader = PdfReader(file)
            fields = reader.get_form_text_fields()
            return [field.strip() for field in fields.keys()] if fields else []
        except Exception as e:
            logger.error(f"Error reading PDF header: {e}")
            return []

    @staticmethod
    def get_excel_header(file):
        try:
            file.seek(0)

            file_name = getattr(file, 'name', '')
            file_extension = os.path.splitext(file_name)[1].lower() if file_name else ''

            if file_extension == '.xls':
                try:
                    import xlrd
                    workbook = xlrd.open_workbook(file_contents=file.read())
                    sheet = workbook.sheet_by_index(0)

                    header_row = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
                    return header_row
                except ImportError:
                    logger.error("xlrd library not installed. Cannot process .xls files.")
                    file.seek(0)
                    df = pd.read_excel(file, engine='xlrd', nrows=0)
                    return [col.strip() for col in df.columns]
            else:
                df = pd.read_excel(file, nrows=0)
                return [col.strip() for col in df.columns]

        except Exception as e:
            logger.error(f"Error reading Excel header: {e}")
            return []

    @staticmethod
    def get_json_header(file):
        try:
            file.seek(0)
            data = json.load(file)
            if isinstance(data, list) and len(data) > 0:
                return [key.strip() for key in data[0].keys()]
            elif isinstance(data, dict):
                return [key.strip() for key in data.keys()]
            return []
        except Exception as e:
            logger.error(f"Error reading JSON header: {e}")
            return []

    @staticmethod
    def get_xml_header(file):
        try:
            file.seek(0)
            tree = ET.parse(file)
            root = tree.getroot()
            return [elem.tag.strip() for elem in root[0]] if len(root) > 0 else []
        except Exception as e:
            logger.error(f"Error reading XML header: {e}")
            return []

    @staticmethod
    def get_header(file, file_type):
        """
        Get headers from file based on file type
        """
        try:
            if not file_type:
                file_name = getattr(file, 'name', '')
                if file_name:
                    file_extension = os.path.splitext(file_name)[1].lower()
                    if file_extension in ['.csv', '.tsv', '.txt']:
                        file_type = 'CSV_C_UTF-8'
                    elif file_extension in ['.xlsx', '.xls']:
                        file_type = 'EXCEL'
                    elif file_extension == '.json':
                        file_type = 'JSON'
                    elif file_extension == '.xml':
                        file_type = 'XML'
                    elif file_extension == '.pdf':
                        file_type = 'PDF'
                    else:
                        logger.error(f"Cannot determine file type from extension: {file_extension}")
                        return []

            format_details = None
            if hasattr(FileFormatMapper, 'get_format_details'):
                format_details = FileFormatMapper.get_format_details(file_type)

            if file_type.startswith("CSV"):
                delimiter = format_details.get('delimiter', ',') if format_details else ','
                encoding = format_details.get('encoding', 'utf-8') if format_details else 'utf-8'

                return ProcessHeader.get_csv_header(file, delimiter=delimiter, encoding=encoding)
            elif file_type == "EXCEL":
                return ProcessHeader.get_excel_header(file)
            elif file_type == "PDF":
                return ProcessHeader.get_pdf_header(file)
            elif file_type == "JSON":
                return ProcessHeader.get_json_header(file)
            elif file_type == "XML":
                return ProcessHeader.get_xml_header(file)
            else:
                logger.error(f"Unsupported file type: {file_type}")
                return []
        except Exception as e:
            logger.error(f"Error in get_header: {e}")
            return []
